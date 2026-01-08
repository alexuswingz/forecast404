"""Compare AC (weekly_units_needed) values between our calculation and Excel"""
from openpyxl import load_workbook
from datetime import date, datetime, timedelta

print("Loading Excel file...")
wb = load_workbook('V2.2 AutoForecast 1000 Bananas 2026.1.7 (1).xlsx', data_only=True)

forecast_ws = wb['forecast_18m+']

# Get AC values (column 29) for all future weeks
print("=== Excel AC Values (Column 29) ===")
today = date(2026, 1, 8)
lead_time_end = today + timedelta(days=130)

excel_ac_sum = 0
excel_p_values = []

for row in range(3, forecast_ws.max_row + 1):
    a_val = forecast_ws.cell(row=row, column=1).value
    ac_val = forecast_ws.cell(row=row, column=29).value
    p_val = forecast_ws.cell(row=row, column=16).value
    
    if a_val:
        if isinstance(a_val, datetime):
            a_date = a_val.date()
        elif isinstance(a_val, date):
            a_date = a_val
        else:
            continue
        
        if ac_val and ac_val > 0:
            excel_ac_sum += ac_val
            p_str = f"{p_val:.2f}" if p_val else "0"
            print(f"  Row {row}: {a_date}, P={p_str}, AC={ac_val:.2f}")
            excel_p_values.append((a_date, p_val, ac_val))

print(f"\nExcel Sum of AC: {excel_ac_sum:.1f}")

# Get AD and AE values
ad_val = forecast_ws.cell(row=3, column=30).value  # Column AD
ae_val = forecast_ws.cell(row=3, column=31).value  # Column AE

print(f"\nExcel AD3 (total units needed): {ad_val}")
print(f"Excel AE3 (units to make): {ae_val}")

# Check inventory
inv_ws = wb['Inventory']
inv_total = inv_ws.cell(row=2, column=1).value
print(f"Excel Inventory A2: {inv_total}")

print(f"\nVerification: AD - Inventory = {ad_val - inv_total if ad_val and inv_total else 'N/A'}")

wb.close()
