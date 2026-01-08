"""
Data Import Module - Import data from Excel files into the database

Handles importing:
- Products from Units_Sold sheet
- Weekly sales data
- Inventory data from FBAInventory and AWDInventory
- Vine claims
- Seasonality data
"""

import openpyxl
from datetime import datetime
from typing import Dict, List, Optional
import os

from . import db
from .models import Product, UnitsSold, Inventory, VineClaim, Seasonality, ForecastSettings


def parse_date(value) -> Optional[datetime]:
    """Parse various date formats from Excel"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%m/%d/%Y']:
            try:
                return datetime.strptime(value.split()[0], fmt.split()[0])
            except ValueError:
                continue
    return None


def import_excel_data(excel_path: str) -> Dict:
    """
    Import all data from the AutoForecast Excel file.
    
    Returns a summary of imported data.
    """
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    results = {
        'products': 0,
        'units_sold_records': 0,
        'inventory_records': 0,
        'vine_claims': 0,
        'seasonality_weeks': 0,
        'errors': []
    }
    
    try:
        # Import Products and Units Sold
        if 'Units_Sold' in wb.sheetnames:
            products, units = import_units_sold(wb['Units_Sold'])
            results['products'] = products
            results['units_sold_records'] = units
        
        # Import FBA Inventory
        if 'FBAInventory' in wb.sheetnames:
            inv = import_fba_inventory(wb['FBAInventory'])
            results['inventory_records'] += inv
        
        # Import AWD Inventory
        if 'AWDInventory' in wb.sheetnames:
            inv = import_awd_inventory(wb['AWDInventory'])
            results['inventory_records'] += inv
        
        # Import Vine Claims
        if 'vine_units_claimed' in wb.sheetnames:
            claims = import_vine_claims(wb['vine_units_claimed'])
            results['vine_claims'] = claims
        
        # Import Seasonality
        if 'Seasonality_Input' in wb.sheetnames:
            seasons = import_seasonality(wb['Seasonality_Input'])
            results['seasonality_weeks'] = seasons
        
        db.session.commit()
        
    except Exception as e:
        db.session.rollback()
        results['errors'].append(str(e))
        raise
    finally:
        wb.close()
    
    return results


def import_units_sold(ws) -> tuple:
    """Import products and weekly sales data from Units_Sold sheet"""
    products_created = 0
    records_created = 0
    
    # Get headers (first row has week dates starting from column D)
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        headers.append(val)
    
    # Parse week dates from headers (starting at column D = index 3)
    week_dates = []
    for i, h in enumerate(headers[3:], start=3):
        dt = parse_date(h)
        if dt:
            week_dates.append((i, dt))
    
    # First pass: Create all products
    product_map = {}  # asin -> product
    for row in range(2, ws.max_row + 1):
        asin = ws.cell(row=row, column=1).value
        product_name = ws.cell(row=row, column=2).value
        size = ws.cell(row=row, column=3).value
        
        if not asin:
            continue
        
        asin_str = str(asin)
        product = Product.query.filter_by(asin=asin_str).first()
        if not product:
            product = Product(
                asin=asin_str,
                product_name=str(product_name) if product_name else None,
                size=str(size) if size else None
            )
            db.session.add(product)
            products_created += 1
        product_map[asin_str] = product
    
    db.session.commit()
    
    # Refresh product map with IDs
    for asin in product_map:
        product_map[asin] = Product.query.filter_by(asin=asin).first()
    
    # Clear existing sales data for faster import
    UnitsSold.query.delete()
    db.session.commit()
    
    # Second pass: Bulk insert all sales records
    batch = []
    batch_size = 1000
    
    for row in range(2, ws.max_row + 1):
        asin = ws.cell(row=row, column=1).value
        if not asin:
            continue
        
        product = product_map.get(str(asin))
        if not product:
            continue
        
        for col, week_date in week_dates:
            units_val = ws.cell(row=row, column=col).value
            units = int(float(units_val)) if units_val and str(units_val).replace('.', '').isdigit() else 0
            week_num = week_date.isocalendar()[1]
            
            batch.append(UnitsSold(
                product_id=product.id,
                week_end=week_date.date(),
                week_number=week_num,
                units=units
            ))
            records_created += 1
            
            # Commit in batches
            if len(batch) >= batch_size:
                db.session.bulk_save_objects(batch)
                db.session.commit()
                batch = []
    
    # Commit remaining
    if batch:
        db.session.bulk_save_objects(batch)
        db.session.commit()
    
    return products_created, records_created


def import_fba_inventory(ws) -> int:
    """Import FBA inventory data"""
    records_updated = 0
    
    # Headers are in row 1
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            headers[val.lower().replace('-', '_').replace(' ', '_')] = col
    
    # Process inventory rows
    for row in range(2, ws.max_row + 1):
        asin_col = headers.get('asin', 4)
        asin = ws.cell(row=row, column=asin_col).value
        
        if not asin:
            continue
        
        # Find product
        product = Product.query.filter_by(asin=str(asin)).first()
        if not product:
            # Create product if not exists
            product_name = ws.cell(row=row, column=headers.get('product_name', 5)).value
            product = Product(asin=str(asin), product_name=str(product_name) if product_name else None)
            db.session.add(product)
            db.session.flush()
        
        # Get snapshot date
        snapshot_col = headers.get('snapshot_date', 1)
        snapshot_date = parse_date(ws.cell(row=row, column=snapshot_col).value)
        if not snapshot_date:
            snapshot_date = datetime.now()
        
        # Get or update inventory record
        inventory = Inventory.query.filter_by(
            product_id=product.id,
            snapshot_date=snapshot_date.date()
        ).first()
        
        if not inventory:
            inventory = Inventory(product_id=product.id, snapshot_date=snapshot_date.date())
            db.session.add(inventory)
        
        # Update FBA values
        def get_int(col_name, default_col):
            col = headers.get(col_name, default_col)
            val = ws.cell(row=row, column=col).value
            return int(float(val)) if val and str(val).replace('.', '').replace('-', '').isdigit() else 0
        
        inventory.fba_available = get_int('available', 7)
        inventory.fba_reserved = get_int('pending_removal_quantity', 8)
        inventory.inv_age_0_90 = get_int('inv_age_0_to_90_days', 9)
        inventory.inv_age_91_180 = get_int('inv_age_91_to_180_days', 10)
        inventory.inv_age_181_270 = get_int('inv_age_181_to_270_days', 11)
        inventory.inv_age_271_365 = get_int('inv_age_271_to_365_days', 12)
        inventory.inv_age_365_plus = get_int('inv_age_365_plus_days', 13)
        
        records_updated += 1
    
    return records_updated


def import_awd_inventory(ws) -> int:
    """Import AWD inventory data"""
    records_updated = 0
    
    # Headers are in row 4 for AWD sheet
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=4, column=col).value
        if val:
            key = val.lower().replace(' ', '_').replace('(', '').replace(')', '')
            headers[key] = col
    
    # Process inventory rows starting from row 5
    for row in range(5, ws.max_row + 1):
        asin_col = headers.get('asin', 4)
        asin = ws.cell(row=row, column=asin_col).value
        
        if not asin:
            continue
        
        # Find product
        product = Product.query.filter_by(asin=str(asin)).first()
        if not product:
            product_name = ws.cell(row=row, column=headers.get('product_name', 1)).value
            product = Product(asin=str(asin), product_name=str(product_name) if product_name else None)
            db.session.add(product)
            db.session.flush()
        
        # Get latest inventory for today
        today = datetime.now().date()
        inventory = Inventory.query.filter_by(
            product_id=product.id,
            snapshot_date=today
        ).first()
        
        if not inventory:
            inventory = Inventory(product_id=product.id, snapshot_date=today)
            db.session.add(inventory)
        
        def get_int(key, default_col):
            col = headers.get(key, default_col)
            val = ws.cell(row=row, column=col).value
            return int(float(val)) if val and str(val).replace('.', '').replace('-', '').isdigit() else 0
        
        inventory.awd_available = get_int('available_in_awd_units', 7)
        inventory.awd_reserved = get_int('reserved_in_awd_units', 9)
        inventory.awd_inbound = get_int('inbound_to_awd_units', 5)
        inventory.awd_outbound_to_fba = get_int('outbound_to_fba_units', 15)
        
        records_updated += 1
    
    return records_updated


def import_vine_claims(ws) -> int:
    """Import Vine program claims"""
    records_created = 0
    
    # Headers in row 1
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            headers[val.lower().replace(' ', '_')] = col
    
    for row in range(2, ws.max_row + 1):
        asin = ws.cell(row=row, column=headers.get('asin', 1)).value
        
        if not asin:
            continue
        
        # Find product
        product = Product.query.filter_by(asin=str(asin)).first()
        if not product:
            product_name = ws.cell(row=row, column=headers.get('product', 2)).value
            product = Product(asin=str(asin), product_name=str(product_name) if product_name else None)
            db.session.add(product)
            db.session.flush()
        
        claim_date = parse_date(ws.cell(row=row, column=headers.get('date', 3)).value)
        units = ws.cell(row=row, column=headers.get('units_claimed', 4)).value
        status = ws.cell(row=row, column=headers.get('vine_status', 5)).value
        
        if claim_date:
            claim = VineClaim(
                product_id=product.id,
                claim_date=claim_date.date(),
                units_claimed=int(float(units)) if units else 0,
                status=str(status) if status else None
            )
            db.session.add(claim)
            records_created += 1
    
    return records_created


def import_seasonality(ws) -> int:
    """Import seasonality data from search volume"""
    from .algorithms import calculate_seasonality
    
    records_created = 0
    
    # Get search volume data (column B starting from row 4)
    search_volumes = []
    for row in range(4, min(ws.max_row + 1, 60)):  # Max 56 weeks
        vol = ws.cell(row=row, column=2).value
        if vol and str(vol).replace('.', '').replace('-', '').isdigit():
            search_volumes.append(float(vol))
    
    if search_volumes:
        # Calculate seasonality using our algorithm
        seasonality_data = calculate_seasonality(search_volumes)
        
        # Clear existing and insert new
        Seasonality.query.delete()
        
        for s in seasonality_data:
            record = Seasonality(
                week_of_year=s['week_of_year'],
                search_volume=s['search_volume'],
                sv_peak_env=s['sv_peak_env'],
                sv_peak_env_offset=s['sv_peak_env_offset'],
                sv_smooth_env=s['sv_smooth_env'],
                sv_final_curve=s['sv_final_curve'],
                sv_smooth=s['sv_smooth'],
                sv_smooth_env_final=s['sv_smooth_env_final'],
                seasonality_index=s['seasonality_index'],
                seasonality_multiplier=s['seasonality_multiplier']
            )
            db.session.add(record)
            records_created += 1
    
    return records_created


def init_default_settings():
    """Initialize default forecast settings - matches Excel Settings sheet"""
    defaults = [
        # Key calibration settings for each algorithm
        ('forecast_multiplier', 3.1, 'Calibration factor for 0-6m algorithm (most aggressive)'),
        ('forecast_multiplier_6_18m', 0.4, 'Calibration factor for 6-18m algorithm (most conservative)'),
        ('forecast_multiplier_18m', 1.4, 'Calibration factor for 18m+ algorithm (moderate)'),
        
        # From Excel Settings sheet
        ('amazon_doi_goal', 120, 'Days of inventory coverage goal'),
        ('inbound_lead_time', 30, 'Shipping time to Amazon (days)'),
        ('manufacture_lead_time', 7, 'Production time (days)'),
        ('market_adjustment', 0.05, 'Year-over-year growth factor (5%)'),
        ('velocity_weight', 0.15, 'Weight for velocity adjustment (15%)'),
        
        # Timezone setting
        ('timezone_offset', -7, 'GMT offset for date calculations (e.g., -7 for Pacific Time)'),
        
        # Legacy settings
        ('velocity_adj_factor', 0.1, 'Velocity adjustment factor for 18m+ forecast'),
        ('growth_factor', 0.05, 'Year-over-year growth factor'),
        ('lead_time_days', 90, 'Production lead time in days'),
        ('safety_stock_weeks', 4, 'Safety stock in weeks of coverage'),
        ('smoothing_factor', 0.85, 'Smoothing factor for conservative estimates'),
    ]
    
    for name, value, description in defaults:
        setting = ForecastSettings.query.filter_by(name=name).first()
        if not setting:
            setting = ForecastSettings(name=name, value=value, description=description)
            db.session.add(setting)
    
    db.session.commit()

