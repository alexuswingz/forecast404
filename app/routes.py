"""
Flask Routes - API endpoints and web interface
"""

from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash
from datetime import datetime, date, timedelta
import os

from . import db
from .models import Product, UnitsSold, Inventory, VineClaim, Seasonality, ForecastSettings, ForecastResult
from .algorithms import (
    calculate_seasonality,
    generate_full_forecast,
    calculate_forecast_0_6m,
    calculate_forecast_6_18m,
    calculate_forecast_18m_plus
)
from .data_import import import_excel_data, init_default_settings

# Blueprints
main_bp = Blueprint('main', __name__)
api_bp = Blueprint('api', __name__)


# =============================================================================
# WEB INTERFACE ROUTES
# =============================================================================

@main_bp.route('/')
def index():
    """Dashboard home page with forecast summary"""
    from datetime import date
    
    # Get sort and filter params
    sort_by = request.args.get('sort', 'units_to_make')  # units_to_make, doi_total, inventory
    sort_order = request.args.get('order', 'desc')  # asc, desc
    filter_algo = request.args.get('algo', 'all')  # all, 0-6m, 6-18m, 18m+
    
    # Get products with inventory (active products)
    products_with_inventory = db.session.query(Product, Inventory)\
        .join(Inventory, Product.id == Inventory.product_id)\
        .order_by(Product.product_name).all()
    
    # Get seasonality and settings for forecasts
    seasonality = get_seasonality_data()
    settings = get_forecast_settings()
    today = date.today()
    
    # Calculate forecasts for each product
    product_forecasts = []
    total_units_to_make = 0
    total_inventory = 0
    critical_count = 0
    low_stock_count = 0
    
    for product, inventory in products_with_inventory:
        # Get sales data
        sales_data = get_product_sales_data(product.id)
        inv_dict = {
            'total_inventory': inventory.total_inventory,
            'fba_available': inventory.fba_available
        }
        
        # Calculate product age in months from first sale
        product_age_months = 0
        if sales_data:
            first_sale = min(s.get('week_end') for s in sales_data if s.get('week_end'))
            if first_sale:
                if isinstance(first_sale, str):
                    from datetime import datetime
                    first_sale = datetime.strptime(first_sale, '%Y-%m-%d').date()
                days_old = (today - first_sale).days
                product_age_months = days_old / 30.44  # Average days per month
        
        # Determine which algorithm to use based on age
        if product_age_months < 6:
            active_algo_name = '0-6m'
        elif product_age_months < 18:
            active_algo_name = '6-18m'
        else:
            active_algo_name = '18m+'
        
        # Calculate forecast
        try:
            forecast = generate_full_forecast(
                product.asin,
                sales_data,
                seasonality,
                inv_dict,
                settings,
                today
            )
            
            # Get the appropriate algorithm based on product age
            active_algo = forecast['algorithms'].get(active_algo_name, {})
            units_to_make = active_algo.get('units_to_make', 0)
            doi_total = active_algo.get('doi_total_days', 0)
            doi_fba = active_algo.get('doi_fba_days', 0)
            
            # Determine status based on DOI
            if doi_total < 30:
                status = 'critical'
                critical_count += 1
            elif doi_total < 60:
                status = 'low'
                low_stock_count += 1
            else:
                status = 'good'
            
            product_forecasts.append({
                'product': product,
                'inventory': inventory,
                'units_to_make': int(units_to_make),
                'doi_total': int(doi_total),
                'doi_fba': int(doi_fba),
                'status': status,
                'algorithm': active_algo_name,
                'age_months': int(product_age_months)
            })
            
            total_units_to_make += units_to_make
            total_inventory += inventory.total_inventory
            
        except Exception as e:
            # Skip products with forecast errors
            product_forecasts.append({
                'product': product,
                'inventory': inventory,
                'units_to_make': 0,
                'doi_total': 0,
                'doi_fba': 0,
                'status': 'unknown',
                'algorithm': active_algo_name,
                'age_months': int(product_age_months)
            })
    
    # Filter by algorithm if specified
    if filter_algo != 'all':
        product_forecasts = [p for p in product_forecasts if p['algorithm'] == filter_algo]
    
    # Sort based on parameter
    reverse_order = sort_order == 'desc'
    if sort_by == 'doi_total':
        product_forecasts.sort(key=lambda x: x['doi_total'], reverse=reverse_order)
    elif sort_by == 'inventory':
        product_forecasts.sort(key=lambda x: x['inventory'].total_inventory, reverse=reverse_order)
    elif sort_by == 'age':
        product_forecasts.sort(key=lambda x: x['age_months'], reverse=reverse_order)
    else:  # default: units_to_make
        product_forecasts.sort(key=lambda x: x['units_to_make'], reverse=reverse_order)
    
    # Get summary stats
    total_products = len(products_with_inventory)
    
    # Get FBA and AWD totals
    latest_inventory = db.session.query(
        db.func.sum(Inventory.fba_available),
        db.func.sum(Inventory.awd_available)
    ).first()
    
    fba_total = latest_inventory[0] or 0
    awd_total = latest_inventory[1] or 0
    
    # Count by algorithm
    algo_counts = {'0-6m': 0, '6-18m': 0, '18m+': 0}
    for p in products_with_inventory:
        product = p[0]
        sales_data = get_product_sales_data(product.id)
        if sales_data:
            first_sale = min(s.get('week_end') for s in sales_data if s.get('week_end'))
            if first_sale:
                if isinstance(first_sale, str):
                    first_sale = datetime.strptime(first_sale, '%Y-%m-%d').date()
                days_old = (today - first_sale).days
                age_months = days_old / 30.44
                if age_months < 6:
                    algo_counts['0-6m'] += 1
                elif age_months < 18:
                    algo_counts['6-18m'] += 1
                else:
                    algo_counts['18m+'] += 1
    
    return render_template('index.html',
        product_forecasts=product_forecasts,
        total_products=total_products,
        total_units_to_make=int(total_units_to_make),
        total_inventory=int(total_inventory),
        fba_total=int(fba_total),
        awd_total=int(awd_total),
        critical_count=critical_count,
        low_stock_count=low_stock_count,
        sort_by=sort_by,
        sort_order=sort_order,
        filter_algo=filter_algo,
        algo_counts=algo_counts
    )


@main_bp.route('/product/<asin>')
def product_detail(asin):
    """Product detail page with forecast"""
    product = Product.query.filter_by(asin=asin).first_or_404()
    
    # Get sales history
    sales = UnitsSold.query.filter_by(product_id=product.id)\
        .order_by(UnitsSold.week_end.desc()).limit(52).all()
    
    # Get inventory
    inventory = Inventory.query.filter_by(product_id=product.id)\
        .order_by(Inventory.snapshot_date.desc()).first()
    
    # Get vine claims
    vine_claims = VineClaim.query.filter_by(product_id=product.id).all()
    
    return render_template('product.html',
        product=product,
        sales=sales,
        inventory=inventory,
        vine_claims=vine_claims
    )


@main_bp.route('/import', methods=['GET', 'POST'])
def import_data():
    """Import data from Excel file"""
    import tempfile
    
    if request.method == 'POST':
        excel_file = request.files.get('excel_file')
        if excel_file:
            # Save to temp file (works on Railway's ephemeral filesystem)
            try:
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                    excel_file.save(tmp.name)
                    file_path = tmp.name
                
                results = import_excel_data(file_path)
                init_default_settings()
                
                # Clean up temp file
                os.unlink(file_path)
                
                flash(f"Import successful! Products: {results['products']}, "
                      f"Sales records: {results['units_sold_records']}, "
                      f"Inventory: {results['inventory_records']}", 'success')
            except Exception as e:
                flash(f"Import error: {str(e)}", 'error')
            
            return redirect(url_for('main.index'))
    
    return render_template('import.html')


@main_bp.route('/forecast/<asin>')
def forecast_view(asin):
    """View forecast for a product"""
    from datetime import date as date_type
    
    product = Product.query.filter_by(asin=asin).first_or_404()
    
    # Support custom calculation date (for matching Excel)
    calc_date_str = request.args.get('date')
    if calc_date_str:
        try:
            calc_date = datetime.strptime(calc_date_str, '%Y-%m-%d').date()
        except:
            calc_date = date_type.today()
    else:
        calc_date = date_type.today()
    
    # Get data for forecast
    sales_data = get_product_sales_data(product.id)
    seasonality = get_seasonality_data()
    inventory = get_product_inventory(product.id)
    settings = get_forecast_settings()
    
    # Generate forecast
    forecast = generate_full_forecast(
        product.asin,
        sales_data,
        seasonality,
        inventory,
        settings,
        calc_date
    )
    
    return render_template('forecast.html',
        product=product,
        forecast=forecast,
        calc_date=calc_date.isoformat()
    )


@main_bp.route('/settings', methods=['GET', 'POST'])
def settings():
    """Forecast settings page"""
    if request.method == 'POST':
        for key in request.form:
            if key.startswith('setting_'):
                name = key.replace('setting_', '')
                value = float(request.form[key])
                setting = ForecastSettings.query.filter_by(name=name).first()
                if setting:
                    setting.value = value
        db.session.commit()
        flash('Settings updated', 'success')
        return redirect(url_for('main.settings'))
    
    settings_list = ForecastSettings.query.all()
    return render_template('settings.html', settings=settings_list)


# =============================================================================
# API ENDPOINTS
# =============================================================================

@api_bp.route('/products')
def api_products():
    """Get all products"""
    products = Product.query.all()
    return jsonify([p.to_dict() for p in products])


@api_bp.route('/products/<asin>')
def api_product(asin):
    """Get single product"""
    product = Product.query.filter_by(asin=asin).first_or_404()
    return jsonify(product.to_dict())


@api_bp.route('/products/<asin>/sales')
def api_product_sales(asin):
    """Get sales history for a product"""
    product = Product.query.filter_by(asin=asin).first_or_404()
    
    # Parse date range
    start_date = request.args.get('start')
    end_date = request.args.get('end')
    
    query = UnitsSold.query.filter_by(product_id=product.id)
    
    if start_date:
        query = query.filter(UnitsSold.week_end >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(UnitsSold.week_end <= datetime.strptime(end_date, '%Y-%m-%d').date())
    
    sales = query.order_by(UnitsSold.week_end).all()
    return jsonify([s.to_dict() for s in sales])


@api_bp.route('/products/<asin>/inventory')
def api_product_inventory(asin):
    """Get inventory for a product"""
    product = Product.query.filter_by(asin=asin).first_or_404()
    inventory = Inventory.query.filter_by(product_id=product.id)\
        .order_by(Inventory.snapshot_date.desc()).first()
    return jsonify(inventory.to_dict() if inventory else {})


@api_bp.route('/products/<asin>/forecast')
def api_product_forecast(asin):
    """Generate and return forecast for a product"""
    product = Product.query.filter_by(asin=asin).first_or_404()
    
    # Get forecast type
    forecast_type = request.args.get('type', 'all')  # '0-6m', '6-18m', '18m+', 'all'
    
    # Gather data
    sales_data = get_product_sales_data(product.id)
    prior_year_data = get_prior_year_data(product.id)
    seasonality = get_seasonality_data()
    inventory = get_product_inventory(product.id)
    settings = get_forecast_settings()
    
    # Generate forecast
    forecast = generate_full_forecast(
        product.asin,
        sales_data,
        prior_year_data,
        seasonality,
        inventory,
        settings
    )
    
    if forecast_type != 'all' and forecast_type in forecast['forecasts']:
        return jsonify({
            'product_asin': product.asin,
            'forecast_type': forecast_type,
            'data': forecast['forecasts'][forecast_type],
            'production_needs': forecast['production_needs']
        })
    
    return jsonify(forecast)


@api_bp.route('/seasonality')
def api_seasonality():
    """Get seasonality data"""
    seasonality = Seasonality.query.order_by(Seasonality.week_of_year).all()
    return jsonify([s.to_dict() for s in seasonality])


@api_bp.route('/seasonality', methods=['POST'])
def api_update_seasonality():
    """Update seasonality from new search volume data"""
    data = request.get_json()
    search_volumes = data.get('search_volumes', [])
    
    if not search_volumes:
        return jsonify({'error': 'No search volume data provided'}), 400
    
    # Calculate new seasonality
    seasonality_data = calculate_seasonality(search_volumes)
    
    # Update database
    Seasonality.query.delete()
    for s in seasonality_data:
        record = Seasonality(
            week_of_year=s['week_of_year'],
            search_volume=s['search_volume'],
            sv_peak_env=s['sv_peak_env'],
            sv_peak_env_offset=s['sv_peak_env_offset'],
            sv_smooth_env=s['sv_smooth_env'],
            sv_final_curve=s['sv_final_curve'],
            sv_smooth=s['sv_smooth'],
            sv_smooth_env_final=s['sv_smooth_env_final'],
            seasonality_index=s['seasonality_index'],
            seasonality_multiplier=s['seasonality_multiplier']
        )
        db.session.add(record)
    
    db.session.commit()
    return jsonify({'message': 'Seasonality updated', 'weeks': len(seasonality_data)})


@api_bp.route('/seasonality/upload', methods=['POST'])
def api_upload_seasonality():
    """Upload seasonality data from Excel template file.
    
    Expected format:
    - Column A: Week (1-52)
    - Column B: Seasonality_Data (search volume)
    - Column C: Child_ASIN (optional, for reference)
    """
    from openpyxl import load_workbook
    import tempfile
    
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'File must be an Excel file (.xlsx or .xls)'}), 400
    
    try:
        # Save to temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name
        
        # Read the Excel file
        wb = load_workbook(tmp_path, data_only=True)
        ws = wb.active
        
        # Read search volumes by week number (1-52)
        search_volumes = [0] * 52  # Initialize 52 weeks
        child_asin = None
        weeks_found = 0
        
        # Read data starting from row 2 (row 1 is header)
        for row in range(2, ws.max_row + 1):
            week_val = ws.cell(row=row, column=1).value
            volume = ws.cell(row=row, column=2).value
            asin = ws.cell(row=row, column=3).value
            
            if week_val is not None and volume is not None:
                try:
                    week_num = int(week_val)
                    if 1 <= week_num <= 52:
                        search_volumes[week_num - 1] = float(volume)
                        weeks_found += 1
                        if asin and not child_asin:
                            child_asin = str(asin)
                except (ValueError, TypeError):
                    continue
        
        wb.close()
        
        # Clean up temp file
        import os
        os.unlink(tmp_path)
        
        if weeks_found == 0:
            return jsonify({'error': 'No valid data found in file. Expected: Week (1-52) in Column A, Seasonality_Data in Column B'}), 400
        
        # Handle missing weeks by interpolation
        for i in range(52):
            if search_volumes[i] == 0:
                # Find nearest non-zero values
                prev_val = None
                next_val = None
                for j in range(1, 52):
                    if prev_val is None and search_volumes[(i - j) % 52] > 0:
                        prev_val = search_volumes[(i - j) % 52]
                    if next_val is None and search_volumes[(i + j) % 52] > 0:
                        next_val = search_volumes[(i + j) % 52]
                    if prev_val and next_val:
                        break
                if prev_val and next_val:
                    search_volumes[i] = (prev_val + next_val) / 2
                elif prev_val:
                    search_volumes[i] = prev_val
                elif next_val:
                    search_volumes[i] = next_val
        
        # Calculate full seasonality
        seasonality_data = calculate_full_seasonality(search_volumes)
        
        # Update database
        Seasonality.query.delete()
        for s in seasonality_data:
            record = Seasonality(
                week_of_year=s['week_of_year'],
                search_volume=s['search_volume'],
                sv_peak_env=s['sv_peak_env'],
                sv_peak_env_offset=s['sv_peak_env_offset'],
                sv_smooth_env=s['sv_smooth_env'],
                sv_final_curve=s['sv_final_curve'],
                sv_smooth=s['sv_smooth'],
                sv_smooth_env_final=s['sv_smooth_env_final'],
                seasonality_index=s['seasonality_index'],
                seasonality_multiplier=s['seasonality_multiplier']
            )
            db.session.add(record)
        
        db.session.commit()
        
        peak_week = max(range(52), key=lambda i: seasonality_data[i]['seasonality_index']) + 1
        
        return jsonify({
            'success': True,
            'message': 'Seasonality data imported successfully',
            'weeks': weeks_found,
            'peak_week': peak_week,
            'child_asin': child_asin
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def calculate_full_seasonality(search_volumes):
    """Calculate all seasonality metrics from weekly search volumes."""
    n = len(search_volumes)
    if n == 0:
        return []
    
    # Peak envelope
    sv_peak_env = []
    for i in range(n):
        start = max(0, i - 2)
        end = min(n, i + 1)
        window = search_volumes[start:end]
        sv_peak_env.append(max(window) if window else search_volumes[i])
    
    # Peak envelope offset
    sv_peak_env_offset = []
    for i in range(n):
        if i < n - 1:
            sv_peak_env_offset.append((sv_peak_env[i] + sv_peak_env[i + 1]) / 2)
        else:
            sv_peak_env_offset.append(sv_peak_env[i])
    
    # Smooth envelope
    sv_smooth_env = []
    for i in range(n):
        start = max(0, i - 1)
        end = min(n, i + 2)
        window = sv_peak_env_offset[start:end]
        sv_smooth_env.append(sum(window) / len(window) if window else sv_peak_env_offset[i])
    
    # Final curve
    sv_final_curve = []
    for i in range(n):
        vals = [search_volumes[i], sv_peak_env_offset[i], sv_smooth_env[i]]
        sv_final_curve.append(sum(vals) / len(vals))
    
    # Smooth
    sv_smooth = []
    for i in range(n):
        start = max(0, i - 1)
        end = min(n, i + 2)
        sv_smooth.append(sum(sv_final_curve[start:end]) / len(sv_final_curve[start:end]))
    
    # Final smooth
    sv_smooth_final = []
    for i in range(n):
        if i < n - 1:
            sv_smooth_final.append((sv_smooth[i] + sv_smooth[i + 1]) / 2)
        else:
            sv_smooth_final.append(sv_smooth[i])
    
    max_h = max(sv_smooth_final) if sv_smooth_final else 1
    avg_h = sum(sv_smooth_final) / len(sv_smooth_final) if sv_smooth_final else 1
    
    results = []
    for i in range(n):
        results.append({
            'week_of_year': i + 1,
            'search_volume': sv_smooth_final[i],
            'sv_peak_env': sv_peak_env[i],
            'sv_peak_env_offset': sv_peak_env_offset[i],
            'sv_smooth_env': sv_smooth_env[i],
            'sv_final_curve': sv_final_curve[i],
            'sv_smooth': sv_smooth[i],
            'sv_smooth_env_final': sv_smooth_final[i],
            'seasonality_index': sv_smooth_final[i] / max_h if max_h > 0 else 0,
            'seasonality_multiplier': sv_smooth_final[i] / avg_h if avg_h > 0 else 1
        })
    
    return results


@api_bp.route('/inventory', methods=['POST'])
def api_update_inventory():
    """Update inventory for a product"""
    data = request.get_json()
    asin = data.get('asin')
    
    product = Product.query.filter_by(asin=asin).first()
    if not product:
        return jsonify({'error': 'Product not found'}), 404
    
    # Get or create today's inventory
    today = date.today()
    inventory = Inventory.query.filter_by(
        product_id=product.id,
        snapshot_date=today
    ).first()
    
    if not inventory:
        inventory = Inventory(product_id=product.id, snapshot_date=today)
        db.session.add(inventory)
    
    # Update fields
    if 'fba_available' in data:
        inventory.fba_available = data['fba_available']
    if 'fba_reserved' in data:
        inventory.fba_reserved = data['fba_reserved']
    if 'fba_inbound' in data:
        inventory.fba_inbound = data['fba_inbound']
    if 'awd_available' in data:
        inventory.awd_available = data['awd_available']
    if 'awd_reserved' in data:
        inventory.awd_reserved = data['awd_reserved']
    if 'awd_inbound' in data:
        inventory.awd_inbound = data['awd_inbound']
    if 'awd_outbound_to_fba' in data:
        inventory.awd_outbound_to_fba = data['awd_outbound_to_fba']
    
    db.session.commit()
    return jsonify(inventory.to_dict())


@api_bp.route('/import', methods=['POST'])
def api_import():
    """Import data from Excel file path"""
    data = request.get_json()
    file_path = data.get('file_path')
    
    if not file_path or not os.path.exists(file_path):
        return jsonify({'error': 'File not found'}), 404
    
    try:
        results = import_excel_data(file_path)
        init_default_settings()
        return jsonify(results)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@api_bp.route('/settings')
def api_get_settings():
    """Get forecast settings"""
    settings = ForecastSettings.query.all()
    return jsonify({s.name: {'value': s.value, 'description': s.description} for s in settings})


@api_bp.route('/settings', methods=['POST'])
def api_update_settings():
    """Update forecast settings"""
    data = request.get_json()
    
    for name, value in data.items():
        setting = ForecastSettings.query.filter_by(name=name).first()
        if setting:
            setting.value = float(value)
        else:
            setting = ForecastSettings(name=name, value=float(value))
            db.session.add(setting)
    
    db.session.commit()
    return jsonify({'message': 'Settings updated'})


@api_bp.route('/dashboard/summary')
def api_dashboard_summary():
    """Get dashboard summary data"""
    # Count products
    total_products = Product.query.count()
    
    # Get total inventory across all products
    inventory_totals = db.session.query(
        db.func.sum(Inventory.fba_available).label('fba'),
        db.func.sum(Inventory.awd_available).label('awd'),
        db.func.sum(Inventory.fba_inbound).label('fba_inbound'),
        db.func.sum(Inventory.awd_inbound).label('awd_inbound')
    ).first()
    
    # Get recent sales (last 4 weeks)
    four_weeks_ago = date.today() - timedelta(weeks=4)
    recent_sales = db.session.query(
        db.func.sum(UnitsSold.units)
    ).filter(UnitsSold.week_end >= four_weeks_ago).scalar() or 0
    
    # Get products running low on inventory
    low_inventory_products = []
    products = Product.query.all()
    for p in products:
        inv = Inventory.query.filter_by(product_id=p.id)\
            .order_by(Inventory.snapshot_date.desc()).first()
        if inv and inv.total_inventory < 100:  # Threshold
            low_inventory_products.append({
                'asin': p.asin,
                'name': p.product_name,
                'inventory': inv.total_inventory
            })
    
    return jsonify({
        'total_products': total_products,
        'inventory': {
            'fba_available': inventory_totals.fba or 0,
            'awd_available': inventory_totals.awd or 0,
            'fba_inbound': inventory_totals.fba_inbound or 0,
            'awd_inbound': inventory_totals.awd_inbound or 0,
            'total': (inventory_totals.fba or 0) + (inventory_totals.awd or 0)
        },
        'recent_sales_4w': recent_sales,
        'low_inventory_products': low_inventory_products[:10]
    })


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def get_product_sales_data(product_id: int) -> list:
    """Get sales data formatted for forecast algorithms"""
    sales = UnitsSold.query.filter_by(product_id=product_id)\
        .order_by(UnitsSold.week_end).all()
    return [
        {
            'week_end': s.week_end.isoformat(),
            'week_number': s.week_number,
            'units': s.units
        }
        for s in sales
    ]


def get_prior_year_data(product_id: int) -> list:
    """Get prior year sales data"""
    one_year_ago = date.today() - timedelta(days=365)
    two_years_ago = date.today() - timedelta(days=730)
    
    sales = UnitsSold.query.filter_by(product_id=product_id)\
        .filter(UnitsSold.week_end >= two_years_ago)\
        .filter(UnitsSold.week_end < one_year_ago)\
        .order_by(UnitsSold.week_end).all()
    
    return [
        {
            'week_end': s.week_end.isoformat(),
            'week_number': s.week_number,
            'units': s.units
        }
        for s in sales
    ]


def get_seasonality_data() -> list:
    """Get seasonality data formatted for algorithms"""
    seasonality = Seasonality.query.order_by(Seasonality.week_of_year).all()
    return [s.to_dict() for s in seasonality]


def get_product_inventory(product_id: int) -> dict:
    """Get current inventory for a product"""
    inventory = Inventory.query.filter_by(product_id=product_id)\
        .order_by(Inventory.snapshot_date.desc()).first()
    
    if inventory:
        return {
            'total_inventory': inventory.total_inventory,
            'fba_available': inventory.fba_available or 0,
            'fba_reserved': inventory.fba_reserved or 0,
            'fba_inbound': inventory.fba_inbound or 0,
            'awd_available': inventory.awd_available or 0,
            'awd_reserved': inventory.awd_reserved or 0,
            'awd_inbound': inventory.awd_inbound or 0
        }
    return {
        'total_inventory': 0,
        'fba_available': 0,
        'fba_reserved': 0,
        'fba_inbound': 0,
        'awd_available': 0,
        'awd_reserved': 0,
        'awd_inbound': 0
    }


def get_forecast_settings() -> dict:
    """Get forecast settings as dictionary"""
    settings = ForecastSettings.query.all()
    result = {s.name: s.value for s in settings}
    
    # Add defaults if not present
    defaults = {
        'amazon_doi_goal': 120,
        'inbound_lead_time': 30,
        'manufacture_lead_time': 7,
        'market_adjustment': 0.05,
        'velocity_adjustment': 24.0,
        'velocity_weight': 0.15,
        'forecast_multiplier': 3.1,  # Calibration factor for 0-6m algorithm (aggressive)
        'forecast_multiplier_6_18m': 0.4,  # Calibration factor for 6-18m algorithm (conservative)
        'forecast_multiplier_18m': 1.4,  # Calibration factor for 18m+ algorithm (moderate)
        'timezone_offset': -7  # GMT-7 (Pacific Time) - matches Excel timezone
    }
    for k, v in defaults.items():
        if k not in result:
            result[k] = v
    
    return result

