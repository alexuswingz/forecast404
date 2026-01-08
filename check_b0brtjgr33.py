"""Check B0BRTJGR33 values from Excel."""

import openpyxl

wb = openpyxl.load_workbook('V2.2 AutoForecast 1000 Bananas 2026.1.7 (1).xlsx', data_only=True)

# Check Settings for this ASIN
ws_settings = wb['Settings']
print("Settings sheet - Algorithm outputs for B0BRTJGR33:")
print("="*60)

# Find rows with Units to Make, DOI values
for row in range(1, 80):
    a = ws_settings.cell(row=row, column=1).value
    b = ws_settings.cell(row=row, column=2).value
    if a:
        a_str = str(a).lower()
        if 'units' in a_str or 'doi' in a_str or 'algorithm' in a_str or 'asin' in a_str:
            print(f"Row {row}: {a} = {b}")

# Check Inventory sheet
ws_inv = wb['Inventory']
print("\n" + "="*60)
print("Inventory values:")
print(f"  A2 (Total): {ws_inv['A2'].value}")
print(f"  B2 (FBA): {ws_inv['B2'].value}")

# Check forecast_18m+ sheet key cells
ws_fc = wb['forecast_18m+']
print("\n" + "="*60)
print("forecast_18m+ sheet key outputs:")
print(f"  AE3 (Units to Make): {ws_fc.cell(row=3, column=31).value}")
print(f"  V3 (DOI Total): {ws_fc.cell(row=3, column=22).value}")
print(f"  AB3 (DOI FBA): {ws_fc.cell(row=3, column=28).value}")
print(f"  AD3 (Total Needed): {ws_fc.cell(row=3, column=30).value}")

wb.close()

