"""Re-import sales data from Excel for all ASINs"""
from openpyxl import load_workbook
from app import create_app, db
from app.models import Product, UnitsSold
from datetime import datetime, date

print("Loading Excel...")
wb = load_workbook('V2.2 AutoForecast 1000 Bananas 2026.1.7 (1).xlsx', data_only=True)
us_ws = wb['Units_Sold']

# Get column headers (dates) starting from column D (4)
excel_dates = []
for col in range(4, us_ws.max_column + 1):
    header = us_ws.cell(row=1, column=col).value
    if header:
        if isinstance(header, datetime):
            excel_dates.append((col, header.date()))
        elif isinstance(header, date):
            excel_dates.append((col, header))

print(f"Found {len(excel_dates)} week columns")
print(f"Date range: {excel_dates[0][1]} to {excel_dates[-1][1]}")

# Read all ASIN data
asin_data = {}
for row in range(2, us_ws.max_row + 1):
    asin = us_ws.cell(row=row, column=1).value
    if not asin or not isinstance(asin, str) or len(asin) != 10:
        continue
    
    sales = {}
    for col, week_date in excel_dates:
        units = us_ws.cell(row=row, column=col).value
        if units is not None:
            try:
                sales[week_date] = int(units)
            except:
                sales[week_date] = 0
        else:
            sales[week_date] = 0
    
    asin_data[asin] = sales

print(f"Found sales data for {len(asin_data)} ASINs")
wb.close()

# Update database
app = create_app()
with app.app_context():
    updated_asins = 0
    updated_records = 0
    
    for asin, sales in asin_data.items():
        product = Product.query.filter_by(asin=asin).first()
        if not product:
            continue
        
        # Delete existing sales
        UnitsSold.query.filter_by(product_id=product.id).delete()
        
        # Insert new sales
        for week_date, units in sales.items():
            record = UnitsSold(
                product_id=product.id,
                week_end=week_date,
                units=units
            )
            db.session.add(record)
            updated_records += 1
        
        updated_asins += 1
    
    db.session.commit()
    print(f"\nUpdated {updated_asins} ASINs with {updated_records} sales records")
    
    # Verify a sample
    print("\n=== Verification for B0C73TDZCQ ===")
    product = Product.query.filter_by(asin='B0C73TDZCQ').first()
    if product:
        sales = UnitsSold.query.filter_by(product_id=product.id).order_by(UnitsSold.week_end).all()
        
        # Show Jan 2025 data
        print("Jan 2025:")
        for s in sales:
            if s.week_end.year == 2025 and s.week_end.month == 1:
                excel_val = asin_data.get('B0C73TDZCQ', {}).get(s.week_end, 'N/A')
                print(f"  {s.week_end}: DB={s.units}, Excel={excel_val}")
