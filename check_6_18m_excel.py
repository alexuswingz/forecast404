"""Check the 6-18m forecast values directly from Excel."""

import openpyxl
from datetime import date

wb = openpyxl.load_workbook('V2.2 AutoForecast 1000 Bananas 2026.1.7 (1).xlsx', data_only=True)
ws = wb['forecast_6m-18m_V2']

print("Column headers (row 2):")
for col in range(1, 30):
    val = ws.cell(row=2, column=col).value
    if val:
        print(f"  Col {col}: {val}")

print("\n" + "="*80)
print("Excel 6-18m forecast data for future weeks:")
print("="*80)

TODAY = date(2026, 1, 8)

print(f"\n{'Row':<5} {'Week End':<12} {'C:units':<8} {'D:season':<10} {'E:deseas':<10} {'F:const':<10} {'G:mult':<8} {'H:adj':<10} {'I:final':<10} {'J:forecast':<10}")
print("-" * 110)

count = 0
for row in range(3, 150):
    week_end = ws.cell(row=row, column=1).value
    if week_end:
        if hasattr(week_end, 'date'):
            week_end = week_end.date()
        
        if week_end >= TODAY:
            c = ws.cell(row=row, column=3).value
            d = ws.cell(row=row, column=4).value
            e = ws.cell(row=row, column=5).value
            f = ws.cell(row=row, column=6).value
            g = ws.cell(row=row, column=7).value
            h = ws.cell(row=row, column=8).value
            i = ws.cell(row=row, column=9).value
            j = ws.cell(row=row, column=10).value
            
            def fmt(v):
                if v is None:
                    return "None"
                elif isinstance(v, (int, float)):
                    return f"{v:.2f}"
                return str(v)[:8]
            
            print(f"{row:<5} {str(week_end):<12} {fmt(c):<8} {fmt(d):<10} {fmt(e):<10} {fmt(f):<10} {fmt(g):<8} {fmt(h):<10} {fmt(i):<10} {fmt(j):<10}")
            count += 1
            if count >= 15:
                break

# Check key cells
print("\n" + "="*60)
print("Key summary cells:")
print("="*60)
print(f"Y3 (Units to Make): {ws.cell(row=3, column=25).value}")
print(f"X3 (Total Needed): {ws.cell(row=3, column=24).value}")
print(f"P3 (DOI Total): {ws.cell(row=3, column=16).value}")
print(f"V3 (DOI FBA): {ws.cell(row=3, column=22).value}")

# Check W column (weekly needed) sum
w_sum = 0
for row in range(3, 200):
    w = ws.cell(row=row, column=23).value
    if isinstance(w, (int, float)):
        w_sum += w
print(f"Sum of W column: {w_sum:.2f}")

wb.close()

