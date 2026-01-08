"""Compare sales data between database and Excel"""
from openpyxl import load_workbook
from app import create_app, db
from app.models import Product, UnitsSold
from datetime import datetime, date

# Load Excel
print("Loading Excel...")
wb = load_workbook('V2.2 AutoForecast 1000 Bananas 2026.1.7 (1).xlsx', data_only=True)
us_ws = wb['Units_Sold']

# Get column headers (dates)
excel_dates = []
for col in range(4, us_ws.max_column + 1):
    header = us_ws.cell(row=1, column=col).value
    if header:
        if isinstance(header, datetime):
            excel_dates.append((col, header.date()))
        elif isinstance(header, date):
            excel_dates.append((col, header))

print(f"Excel has {len(excel_dates)} weeks of data")
print(f"Date range: {excel_dates[0][1]} to {excel_dates[-1][1]}")

# Find B0C73TDZCQ row
asin = 'B0C73TDZCQ'
excel_sales = {}
for row in range(2, us_ws.max_row + 1):
    row_asin = us_ws.cell(row=row, column=1).value
    if row_asin == asin:
        print(f"\nFound {asin} at row {row}")
        for col, week_date in excel_dates:
            units = us_ws.cell(row=row, column=col).value
            if units:
                excel_sales[week_date] = int(units)
        break

print(f"Excel has {len(excel_sales)} non-zero weeks for {asin}")

wb.close()

# Get database sales
app = create_app()
with app.app_context():
    product = Product.query.filter_by(asin=asin).first()
    units_sold = UnitsSold.query.filter_by(product_id=product.id).order_by(UnitsSold.week_end).all()
    
    db_sales = {u.week_end: u.units for u in units_sold}
    print(f"Database has {len(db_sales)} weeks for {asin}")
    
    # Compare
    print("\n=== Comparison ===")
    
    # Get Jan 2025 data (prior year for forecasting Jan 2026)
    print("\nJan 2025 sales (prior year):")
    for week in sorted(excel_sales.keys()):
        if week.year == 2025 and week.month == 1:
            excel_val = excel_sales.get(week, 0)
            db_val = db_sales.get(week, 0)
            match = "OK" if excel_val == db_val else "DIFF"
            print(f"  {week}: Excel={excel_val}, DB={db_val} {match}")
    
    # Check recent data (Dec 2025, Jan 2026)
    print("\nRecent sales (Dec 2025 - Jan 2026):")
    for week in sorted(excel_sales.keys()):
        if (week.year == 2025 and week.month == 12) or (week.year == 2026 and week.month == 1):
            excel_val = excel_sales.get(week, 0)
            db_val = db_sales.get(week, 0)
            match = "OK" if excel_val == db_val else "DIFF"
            print(f"  {week}: Excel={excel_val}, DB={db_val} {match}")
    
    # Count differences
    diff_count = 0
    for week, excel_val in excel_sales.items():
        db_val = db_sales.get(week, 0)
        if excel_val != db_val:
            diff_count += 1
    
    print(f"\nTotal weeks with differences: {diff_count}")
