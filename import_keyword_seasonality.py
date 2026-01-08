"""Import Keyword_Seasonality data from Excel."""

import openpyxl
from app import create_app, db
from app.models import Seasonality

wb = openpyxl.load_workbook('V2.2 AutoForecast 1000 Bananas 2026.1.7 (1).xlsx', data_only=True)
ws = wb['Keyword_Seasonality']

print("Importing Keyword_Seasonality data...")

# Find the data start (row 4 based on earlier check)
data = []
for row in range(4, 60):  # Weeks 1-52 should be in rows 4-55
    week = ws.cell(row=row, column=1).value
    sv_smooth = ws.cell(row=row, column=9).value  # Column I: sv_smooth_env
    season_idx = ws.cell(row=row, column=10).value  # Column J: seasonality_index
    
    if week and isinstance(week, (int, float)):
        week_int = int(week)
        if 1 <= week_int <= 52:
            data.append({
                'week_of_year': week_int,
                'sv_smooth_env': sv_smooth or 0,
                'seasonality_index': season_idx or 1.0
            })

print(f"Found {len(data)} weeks of seasonality data")
print("\nSample data:")
for d in data[:5]:
    print(f"  Week {d['week_of_year']}: sv_smooth={d['sv_smooth_env']:.2f}, index={d['seasonality_index']:.4f}")

wb.close()

# Update database
app = create_app()
with app.app_context():
    # Clear existing seasonality
    Seasonality.query.delete()
    
    # Insert new data
    for d in data:
        # Calculate multiplier from index (index is already 0-1 range)
        # Multiplier should be relative to average
        avg_index = sum(x['seasonality_index'] for x in data) / len(data)
        multiplier = d['seasonality_index'] / avg_index if avg_index > 0 else 1.0
        
        s = Seasonality(
            week_of_year=d['week_of_year'],
            search_volume=d['sv_smooth_env'],  # Store raw search volume
            sv_smooth_env=d['sv_smooth_env'],
            seasonality_index=d['seasonality_index'],
            seasonality_multiplier=multiplier
        )
        db.session.add(s)
    
    db.session.commit()
    print(f"\nImported {len(data)} seasonality records")
    
    # Verify
    print("\nVerification - first 5 weeks:")
    for s in Seasonality.query.order_by(Seasonality.week_of_year).limit(5).all():
        print(f"  Week {s.week_of_year}: sv_smooth={s.search_volume:.2f}, index={s.seasonality_index:.4f}")

